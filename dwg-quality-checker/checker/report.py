"""
Geradores de relatório: console (rich), HTML (jinja2), CSV, PDF e XLSX.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict

from .i18n import _, get_tr_dict
from .rules import Severity


# ─────────────────────────────────────────────────────────────────────────────
#  Console  (rich)
# ─────────────────────────────────────────────────────────────────────────────


def print_console_report(result: Dict) -> None:
    """Imprime o resultado no terminal usando rich."""
    from rich import box
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table

    console = Console()

    color = "green" if result["passed"] else "red"
    status = "✅  APROVADO" if result["passed"] else "❌  REPROVADO"

    console.print()
    console.print(
        Panel(
            f"[bold {color}]{status}[/bold {color}]\n\n"
            f"Arquivo : [cyan]{result['file']}[/cyan]\n"
            f"[red]Erros   : {result['errors']}[/red]   "
            f"[yellow]Avisos : {result['warnings']}[/yellow]   "
            f"[blue]Infos  : {result['infos']}[/blue]",
            title="[bold white]DWG Quality Checker[/bold white]",
            border_style=color,
            padding=(1, 2),
        )
    )

    if not result["issues"]:
        console.print("[bold green]\n  Nenhum problema encontrado! 🎉[/bold green]\n")
        return

    table = Table(
        title="Problemas Encontrados",
        box=box.ROUNDED,
        show_lines=True,
        expand=True,
    )
    table.add_column("Severidade", style="bold", width=12, justify="center")
    table.add_column("Regra", style="cyan", width=34)
    table.add_column("Mensagem")
    table.add_column("Layer", style="dim", width=18)
    table.add_column("Localização", style="dim", width=22)
    table.add_column("Detalhes", style="dim", width=30)

    _colors = {
        Severity.ERROR: "red",
        Severity.WARNING: "yellow",
        Severity.INFO: "blue",
    }

    for issue in result["issues"]:
        c = _colors.get(issue.severity, "white")
        table.add_row(
            f"[{c}]{issue.severity.value}[/{c}]",
            issue.rule,
            issue.message,
            issue.layer or "—",
            getattr(issue, "location", "") or "—",
            issue.details or "—",
        )

    console.print()
    console.print(table)
    console.print()


# ─────────────────────────────────────────────────────────────────────────────
#  HTML  (jinja2)
# ─────────────────────────────────────────────────────────────────────────────


def generate_html_report(result: Dict, output_path: str | None = None) -> str:
    """Gera relatório em HTML usando o template Jinja2."""
    import json
    import sys
    from jinja2 import Environment, FileSystemLoader

    if getattr(sys, 'frozen', False):
        template_dir = Path(sys._MEIPASS) / "templates"
    else:
        template_dir = Path(__file__).parent.parent / "templates"
    env = Environment(loader=FileSystemLoader(str(template_dir)), autoescape=False)
    env.filters["tojson"] = lambda v: json.dumps(v, ensure_ascii=False)
    template = env.get_template("report.html")

    html = template.render(
        result=result,
        Severity=Severity,
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        tr=get_tr_dict(),
    )

    if output_path is None:
        output_path = Path(result["file_path"]).stem + "_report.html"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    return str(output_path)


# ─────────────────────────────────────────────────────────────────────────────
#  PDF  (reportlab)
# ─────────────────────────────────────────────────────────────────────────────


def generate_pdf_report(result: Dict, output_path: str | None = None) -> str:
    """Gera relatório em PDF profissional com logo Vantara Tech."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    if output_path is None:
        output_path = Path(result["file_path"]).stem + "_report.pdf"

    # ── Cores da identidade visual ────────────────────────────────────────────
    C_BG      = colors.HexColor("#0f1117")
    C_ACCENT  = colors.HexColor("#3b82f6")
    C_ERROR   = colors.HexColor("#ef4444")
    C_WARNING = colors.HexColor("#f59e0b")
    C_INFO    = colors.HexColor("#3b82f6")
    C_SUCCESS = colors.HexColor("#22c55e")
    C_TEXT    = colors.HexColor("#e2e8f0")
    C_MUTED   = colors.HexColor("#8892a4")
    C_SURFACE = colors.HexColor("#1a1d27")
    C_ROW_ALT = colors.HexColor("#21263a")

    W, H = A4
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"DWG Quality Checker — {result['file']}",
        author="Vantara Tech — Luiz Q. Melo",
        subject=_("pdf_subject"),
    )

    styles = getSampleStyleSheet()
    def S(name, **kw):
        return ParagraphStyle(name, **kw)

    sTitle   = S("sTitle",   fontSize=18, leading=22, textColor=C_TEXT,    alignment=TA_LEFT,
                              fontName="Helvetica-Bold")
    sSub     = S("sSub",     fontSize=10, leading=13, textColor=C_MUTED,   alignment=TA_LEFT)
    sLabel   = S("sLabel",   fontSize=8,  leading=10, textColor=C_MUTED,   alignment=TA_LEFT,
                              fontName="Helvetica-Bold")
    sValue   = S("sValue",   fontSize=9,  leading=11, textColor=C_TEXT,    alignment=TA_LEFT)
    sMsg     = S("sMsg",     fontSize=8,  leading=10, textColor=C_TEXT,    alignment=TA_LEFT)
    sMuted   = S("sMuted",   fontSize=7,  leading=9,  textColor=C_MUTED,   alignment=TA_LEFT)
    sCenter  = S("sCenter",  fontSize=8,  leading=10, textColor=C_MUTED,   alignment=TA_CENTER)
    sBrand   = S("sBrand",   fontSize=7,  leading=9,  textColor=C_MUTED,   alignment=TA_RIGHT)
    sHash    = S("sHash",    fontSize=6,  leading=8,  textColor=C_MUTED,   alignment=TA_LEFT,
                              fontName="Courier")

    passed   = result["passed"]
    status_c = C_SUCCESS if passed else C_ERROR
    status_t = _("pdf_passed") if passed else _("pdf_failed")
    ts       = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    story = []

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    hdr_data = [[
        Paragraph("🏗 DWG Quality Checker", sTitle),
        Paragraph("Vantara Tech — Luiz Q. Melo", sBrand),
    ]]
    hdr_tbl = Table(hdr_data, colWidths=[W * 0.62, W * 0.27])
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, -1), C_SURFACE),
        ("TOPPADDING",  (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (0, -1), 12),
        ("RIGHTPADDING", (-1, 0), (-1, -1), 12),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 6 * mm))

    # ── Status badge + metadados ──────────────────────────────────────────────
    sStatus = S("sStatus", fontSize=13, leading=16, textColor=status_c,
                fontName="Helvetica-Bold", alignment=TA_LEFT)
    meta_rows = [
        [Paragraph(f"● {status_t}", sStatus), ""],
        [Paragraph(_("pdf_file"),        sLabel), Paragraph(result["file"], sValue)],
        [Paragraph(_("pdf_path"),        sLabel), Paragraph(result.get("file_path",""), sMuted)],
        [Paragraph(_("pdf_size"),        sLabel), Paragraph(f"{result.get('file_size_mb',0):.2f} MB", sValue)],
        [Paragraph(_("pdf_dxf_version"), sLabel), Paragraph(result.get("dxf_version_name","?"), sValue)],
        [Paragraph(_("pdf_entities"),    sLabel), Paragraph(f"{result.get('entity_count',0):,}", sValue)],
        [Paragraph(_("pdf_duration"),    sLabel), Paragraph(f"{result.get('check_time',0):.2f}s", sValue)],
        [Paragraph(_("pdf_checked_at"),  sLabel), Paragraph(ts, sValue)],
    ]
    if result.get("sha256"):
        meta_rows.append([
            Paragraph(_("pdf_sha256"), sLabel),
            Paragraph(result["sha256"], sHash),
        ])

    meta_tbl = Table(meta_rows, colWidths=[32 * mm, W - 32 * mm - 36 * mm])
    meta_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), C_SURFACE),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 10),
        ("SPAN",         (0, 0), (-1, 0)),
        ("TOPPADDING",   (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING",(0, 0), (-1, 0), 8),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 5 * mm))

    # ── Resumo de contagens ───────────────────────────────────────────────────
    sum_data = [[
        Paragraph(f"❌  {result['errors']}", S("E", fontSize=14, fontName="Helvetica-Bold",
                    textColor=C_ERROR, alignment=TA_CENTER)),
        Paragraph(f"⚠  {result['warnings']}", S("W", fontSize=14, fontName="Helvetica-Bold",
                    textColor=C_WARNING, alignment=TA_CENTER)),
        Paragraph(f"ℹ  {result['infos']}", S("I", fontSize=14, fontName="Helvetica-Bold",
                    textColor=C_INFO, alignment=TA_CENTER)),
        Paragraph(f"∑  {result['total_issues']}", S("T", fontSize=14, fontName="Helvetica-Bold",
                    textColor=C_TEXT, alignment=TA_CENTER)),
    ]]
    labels = [[
        Paragraph(_("pdf_errors"),   sCenter),
        Paragraph(_("pdf_warnings"), sCenter),
        Paragraph(_("pdf_infos"),    sCenter),
        Paragraph(_("pdf_total"),    sCenter),
    ]]
    cw = (W - 36 * mm) / 4
    sum_tbl = Table(sum_data + labels, colWidths=[cw] * 4)
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), C_SURFACE),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 6 * mm))

    # ── Tabela de problemas ───────────────────────────────────────────────────
    story.append(Paragraph(_("pdf_issues_title"), S("H2", fontSize=10, fontName="Helvetica-Bold",
                              textColor=C_TEXT, leading=13)))
    story.append(Spacer(1, 2 * mm))

    SEV_COLOR = {"ERROR": C_ERROR, "WARNING": C_WARNING, "INFO": C_INFO}

    if not result["issues"]:
        story.append(Paragraph(_("pdf_no_issues"), S("Ok", fontSize=10,
                                  textColor=C_SUCCESS, leading=13)))
    else:
        col_widths = [18*mm, 40*mm, 70*mm, 28*mm, 12*mm]
        tbl_header = [[
            Paragraph(_("pdf_col_severity"), sLabel),
            Paragraph(_("pdf_col_rule"),     sLabel),
            Paragraph(_("pdf_col_message"),  sLabel),
            Paragraph(_("pdf_col_layer"),    sLabel),
            Paragraph(_("pdf_col_handle"),   sLabel),
        ]]
        tbl_rows = []
        for idx, iss in enumerate(result["issues"]):
            sc = SEV_COLOR.get(iss.severity.value, C_TEXT)
            bg = C_ROW_ALT if idx % 2 else C_SURFACE
            tbl_rows.append([
                Paragraph(iss.severity.value, S(f"sv{idx}", fontSize=7, fontName="Helvetica-Bold",
                            textColor=sc, leading=9)),
                Paragraph(iss.rule or "", sMuted),
                Paragraph((iss.message or "")[:120], sMsg),
                Paragraph(iss.layer or "—", sMuted),
                Paragraph(iss.handle or "—", sMuted),
            ])

        full_table = Table(tbl_header + tbl_rows, colWidths=col_widths, repeatRows=1)
        full_table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#2e3347")),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 5),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_SURFACE, C_ROW_ALT]),
        ]))
        story.append(full_table)

    story.append(Spacer(1, 8 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_MUTED))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        f"DWG Quality Checker v2.5.0  ·  Vantara Tech  ·  Luiz Q. Melo  ·  {ts}",
        sCenter,
    ))

    # ── Construir PDF com fundo escuro ────────────────────────────────────────
    def _dark_bg(canvas, doc):  # noqa: ARG001
        canvas.saveState()
        canvas.setFillColor(C_BG)
        canvas.rect(0, 0, W, H, fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=_dark_bg, onLaterPages=_dark_bg)
    return str(output_path)





def generate_csv_report(result: Dict, output_path: str | None = None) -> str:
    """Gera relatório em CSV (separado por ponto-e-vírgula, UTF-8 BOM)."""
    if output_path is None:
        output_path = Path(result["file_path"]).stem + "_report.csv"

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            _("csv_file"), _("csv_severity"), _("csv_rule"), _("csv_message"),
            _("csv_entity_type"), _("csv_layer"), _("csv_location"),
            _("csv_handle"), _("csv_details"),
        ])
        for issue in result["issues"]:
            writer.writerow(
                [
                    result["file"],
                    issue.severity.value,
                    issue.rule,
                    issue.message,
                    issue.entity_type,
                    issue.layer,
                    getattr(issue, "location", ""),
                    issue.handle,
                    issue.details,
                ]
            )

    return str(output_path)


def generate_excel_report(result: Dict, output_path: str | None = None) -> str | None:
    """Gera relatório Excel (.xlsx) com formatação, abas e cores por severidade.

    Requer openpyxl. Retorna o caminho do arquivo ou None se não instalado.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    if output_path is None:
        output_path = Path(result["file_path"]).stem + "_report.xlsx"

    ts = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # ── Paleta de cores ───────────────────────────────────────────────────────
    C_BG      = "1A1D27"
    C_SURFACE = "21263A"
    C_HEADER  = "2E3347"
    C_TEXT    = "E2E8F0"
    C_MUTED   = "8892A4"
    C_SUCCESS = "22C55E"
    C_ERROR   = "EF4444"
    C_WARNING = "F59E0B"
    C_INFO    = "3B82F6"
    C_ACCENT  = "3B82F6"
    C_ALT     = "1E2235"

    SEV_FILL = {
        "ERROR":   PatternFill("solid", fgColor="3D1515"),
        "WARNING": PatternFill("solid", fgColor="3D2E0A"),
        "INFO":    PatternFill("solid", fgColor="0F2444"),
    }
    SEV_FG = {
        "ERROR":   Font(color=C_ERROR,   bold=True, name="Segoe UI", size=9),
        "WARNING": Font(color=C_WARNING, bold=True, name="Segoe UI", size=9),
        "INFO":    Font(color=C_INFO,    bold=True, name="Segoe UI", size=9),
    }

    def _hdr_font():
        return Font(color=C_TEXT, bold=True, name="Segoe UI", size=9)
    def _muted_font():
        return Font(color=C_MUTED, name="Segoe UI", size=9)
    def _text_font():
        return Font(color=C_TEXT, name="Segoe UI", size=9)
    def _hdr_fill():
        return PatternFill("solid", fgColor=C_HEADER)
    def _alt_fill(idx: int):
        return PatternFill("solid", fgColor=C_ALT if idx % 2 else C_SURFACE)
    def _thin_border():
        s = Side(style="thin", color="2E3347")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # Aba 1 — Resumo / Summary
    # ═══════════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = _("excel_sheet_summary")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 56
    ws_sum.sheet_format.defaultRowHeight = 18

    # Título
    ws_sum["A1"] = "DWG Quality Checker"
    ws_sum["A1"].font = Font(color=C_ACCENT, bold=True, name="Segoe UI", size=14)
    ws_sum["A1"].fill = PatternFill("solid", fgColor=C_BG)

    status_color = C_SUCCESS if result["passed"] else C_ERROR
    status_text  = _("result_passed") if result["passed"] else _("result_failed")
    ws_sum["B1"] = status_text
    ws_sum["B1"].font = Font(color=status_color, bold=True, name="Segoe UI", size=12)
    ws_sum["B1"].fill = PatternFill("solid", fgColor=C_BG)
    ws_sum["B1"].alignment = Alignment(horizontal="right")
    ws_sum.row_dimensions[1].height = 28

    ws_sum.append([])  # linha separadora

    META_ROWS = [
        (_("pdf_file"),         result.get("file", "")),
        (_("pdf_path"),         result.get("file_path", "")),
        (_("pdf_size"),         f"{result.get('file_size_mb', 0):.2f} MB"),
        (_("pdf_dxf_version"),  result.get("dxf_version_name", "—")),
        (_("pdf_entities"),     str(result.get("entity_count", "—"))),
        (_("pdf_checked_at"),   ts),
        (_("pdf_sha256"),       ((result.get("sha256") or "")[:32] + "…") if result.get("sha256") else "—"),
        (_("pdf_duration"),     f"{result.get('check_time', 0):.2f}s"),
    ]
    for label, value in META_ROWS:
        r = ws_sum.max_row + 1
        ws_sum.cell(r, 1, label).font = _muted_font()
        ws_sum.cell(r, 1).fill        = PatternFill("solid", fgColor=C_BG)
        ws_sum.cell(r, 2, value).font = _text_font()
        ws_sum.cell(r, 2).fill        = PatternFill("solid", fgColor=C_BG)

    ws_sum.append([])  # separador

    COUNT_ROWS = [
        (_("pdf_errors"),   result.get("errors",        0), C_ERROR),
        (_("pdf_warnings"), result.get("warnings",      0), C_WARNING),
        (_("pdf_infos"),    result.get("infos",         0), C_INFO),
        (_("pdf_total"),    result.get("total_issues",  0), C_TEXT),
    ]
    for label, value, color in COUNT_ROWS:
        r = ws_sum.max_row + 1
        ws_sum.cell(r, 1, label).font = Font(color=color, bold=True, name="Segoe UI", size=10)
        ws_sum.cell(r, 1).fill        = PatternFill("solid", fgColor=C_SURFACE)
        ws_sum.cell(r, 2, value).font = Font(color=color, bold=True, name="Segoe UI", size=11)
        ws_sum.cell(r, 2).fill        = PatternFill("solid", fgColor=C_SURFACE)
        ws_sum.cell(r, 2).alignment   = Alignment(horizontal="center")
        ws_sum.row_dimensions[r].height = 20

    # ═══════════════════════════════════════════════════════════════════════
    # Aba 2 — Ocorrências / Issues
    # ═══════════════════════════════════════════════════════════════════════
    ws_iss = wb.create_sheet(_("excel_sheet_issues"))
    ws_iss.sheet_view.showGridLines = False
    ws_iss.freeze_panes = "A2"

    ISS_COLS = [
        (_("csv_severity"),    14),
        (_("csv_rule"),        28),
        (_("csv_message"),     52),
        (_("csv_entity_type"), 16),
        (_("csv_layer"),       18),
        (_("csv_location"),    20),
        (_("csv_handle"),      12),
        (_("csv_details"),     40),
    ]
    for col_idx, (title, width) in enumerate(ISS_COLS, start=1):
        cell = ws_iss.cell(1, col_idx, title)
        cell.font      = _hdr_font()
        cell.fill      = _hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws_iss.column_dimensions[get_column_letter(col_idx)].width = width
    ws_iss.row_dimensions[1].height = 22

    for idx, issue in enumerate(result["issues"]):
        row_num = idx + 2
        row_data = [
            issue.severity.value,
            issue.rule or "",
            issue.message or "",
            issue.entity_type or "",
            issue.layer or "—",
            getattr(issue, "location", "") or "—",
            issue.handle or "—",
            issue.details or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_iss.cell(row_num, col_idx, value)
            if col_idx == 1:
                cell.font      = SEV_FG.get(issue.severity.value, _text_font())
                cell.fill      = SEV_FILL.get(issue.severity.value, _alt_fill(idx))
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = _text_font()
                cell.fill = SEV_FILL.get(issue.severity.value, _alt_fill(idx))
            cell.border = _thin_border()
        ws_iss.row_dimensions[row_num].height = 16

    ws_iss.auto_filter.ref = (
        f"A1:{get_column_letter(len(ISS_COLS))}{len(result['issues']) + 1}"
    )

    # ═══════════════════════════════════════════════════════════════════════
    # Aba 3 — Por Layer / By Layer
    # ═══════════════════════════════════════════════════════════════════════
    ws_lyr = wb.create_sheet(_("excel_sheet_bylayer"))
    ws_lyr.sheet_view.showGridLines = False

    layer_map: dict[str, dict] = {}
    for issue in result["issues"]:
        lyr = issue.layer or "—"
        if lyr not in layer_map:
            layer_map[lyr] = {"ERROR": 0, "WARNING": 0, "INFO": 0}
        layer_map[lyr][issue.severity.value] = (
            layer_map[lyr].get(issue.severity.value, 0) + 1
        )

    LYR_COLS = [
        (_("rpt_col_layer"), 22),
        (_("pdf_errors"),    10),
        (_("pdf_warnings"),  12),
        (_("pdf_infos"),     10),
        (_("pdf_total"),     10),
    ]
    for col_idx, (title, width) in enumerate(LYR_COLS, start=1):
        cell = ws_lyr.cell(1, col_idx, title)
        cell.font      = _hdr_font()
        cell.fill      = _hdr_fill()
        cell.alignment = Alignment(horizontal="center")
        ws_lyr.column_dimensions[get_column_letter(col_idx)].width = width
    ws_lyr.row_dimensions[1].height = 22

    for idx, (lyr, counts) in enumerate(
        sorted(layer_map.items(), key=lambda x: -sum(x[1].values()))
    ):
        r = idx + 2
        total = sum(counts.values())
        row_vals = [lyr, counts["ERROR"], counts["WARNING"], counts["INFO"], total]
        for col_idx, val in enumerate(row_vals, start=1):
            cell           = ws_lyr.cell(r, col_idx, val)
            cell.fill      = _alt_fill(idx)
            cell.border    = _thin_border()
            cell.alignment = Alignment(
                horizontal="center" if col_idx > 1 else "left"
            )
            if col_idx == 2 and counts["ERROR"]:
                cell.font = Font(color=C_ERROR,   bold=True, name="Segoe UI", size=9)
            elif col_idx == 3 and counts["WARNING"]:
                cell.font = Font(color=C_WARNING, bold=True, name="Segoe UI", size=9)
            elif col_idx == 4 and counts["INFO"]:
                cell.font = Font(color=C_INFO,    bold=True, name="Segoe UI", size=9)
            else:
                cell.font = _text_font()
        ws_lyr.row_dimensions[r].height = 16

    # ── Salvar ────────────────────────────────────────────────────────────
    wb.save(str(output_path))
    return str(output_path)


# ─────────────────────────────────────────────────────────────────────────────
#  Batch Dashboard HTML
# ─────────────────────────────────────────────────────────────────────────────


def generate_batch_dashboard(results: list, output_path: str | None = None) -> str:
    """Gera HTML de dashboard agregado para múltiplos arquivos verificados em lote.

    Parameters
    ----------
    results:      lista de dicts retornados por DXFChecker.check()
    output_path:  caminho de saída (padrão: batch_dashboard.html)

    Returns
    -------
    str — caminho absoluto do arquivo gerado
    """
    if output_path is None:
        output_path = "batch_dashboard.html"

    ts      = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    total   = len(results)
    passed  = sum(1 for r in results if r["passed"])
    failed  = total - passed
    rate    = (passed / total * 100) if total else 0
    tot_err = sum(r["errors"]   for r in results)
    tot_war = sum(r["warnings"] for r in results)

    # ── Frequência de regras com erros ────────────────────────────────────────
    rule_freq: dict[str, int] = {}
    for r in results:
        for iss in r["issues"]:
            if iss.severity.value == "ERROR":
                rule_freq[iss.rule] = rule_freq.get(iss.rule, 0) + 1
    top_rules = sorted(rule_freq.items(), key=lambda x: -x[1])[:10]

    # ── Tabela de resultados ─────────────────────────────────────────────────
    rows_html = ""
    for r in results:
        icon = "✅" if r["passed"] else "❌"
        cls  = "pass" if r["passed"] else "fail"
        rows_html += (
            f'<tr class="{cls}">'
            f'<td>{r["file"]}</td>'
            f'<td class="center">{icon} {"OK" if r["passed"] else "FALHOU"}</td>'
            f'<td class="center err">{r["errors"]}</td>'
            f'<td class="center warn">{r["warnings"]}</td>'
            f'<td class="center info">{r["infos"]}</td>'
            f'</tr>\n'
        )

    # ── Barras das regras mais frequentes ─────────────────────────────────────
    bar_html = ""
    max_count = top_rules[0][1] if top_rules else 1
    for rule, count in top_rules:
        pct = count / max_count * 100
        bar_html += (
            f'<div class="bar-row">'
            f'<span class="bar-label">{rule}</span>'
            f'<div class="bar-track">'
            f'<div class="bar-fill" style="width:{pct:.0f}%"></div>'
            f'<span class="bar-count">{count}</span>'
            f'</div></div>\n'
        )

    rules_section = (
        f'<section><h2>📊 Erros Mais Frequentes</h2>{bar_html}</section>'
        if bar_html else ""
    )

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8">
  <title>DWG Quality Checker — Dashboard de Lote</title>
  <style>
    :root {{
      --bg:      #0f1117; --surface: #1a1d27; --surface2: #21263a;
      --border:  #2e3347; --text:    #e2e8f0; --muted:    #8892a4;
      --success: #22c55e; --error:   #ef4444; --warning:  #f59e0b;
      --info:    #3b82f6; --accent:  #3b82f6;
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ background: var(--bg); color: var(--text);
            font-family: 'Segoe UI', sans-serif; padding: 24px; }}
    h1 {{ font-size: 1.4rem; color: var(--accent); margin-bottom: 4px; }}
    .sub {{ color: var(--muted); font-size: .85rem; margin-bottom: 20px; }}
    .cards {{ display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }}
    .card {{ background: var(--surface); border: 1px solid var(--border);
             border-radius: 10px; padding: 16px 24px; min-width: 140px; }}
    .card .val {{ font-size: 2rem; font-weight: 700; }}
    .card .lab {{ color: var(--muted); font-size: .8rem; margin-top: 4px; }}
    section {{ background: var(--surface); border: 1px solid var(--border);
               border-radius: 10px; padding: 16px; margin-bottom: 20px; }}
    h2 {{ font-size: 1rem; margin-bottom: 12px; color: var(--text); }}
    table {{ width: 100%; border-collapse: collapse; font-size: .85rem; }}
    th {{ background: var(--surface2); padding: 8px 12px; text-align: left;
          color: var(--muted); font-weight: 600; font-size: .8rem;
          border-bottom: 1px solid var(--border); }}
    td {{ padding: 7px 12px; border-bottom: 1px solid var(--border); }}
    tr.pass td {{ color: var(--text); }}
    tr.fail td {{ color: #fca5a5; }}
    tr:hover td {{ background: var(--surface2); }}
    .center {{ text-align: center; }}
    .err  {{ color: var(--error);   font-weight: 700; }}
    .warn {{ color: var(--warning); font-weight: 700; }}
    .info {{ color: var(--info); }}
    .bar-row {{ display: flex; align-items: center; gap: 10px; margin-bottom: 7px; }}
    .bar-label {{ color: var(--muted); font-size: .78rem; width: 240px;
                  white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    .bar-track {{ flex: 1; background: var(--surface2); border-radius: 4px;
                  height: 18px; position: relative; }}
    .bar-fill  {{ background: var(--error); border-radius: 4px; height: 100%;
                  min-width: 4px; }}
    .bar-count {{ position: absolute; right: 6px; top: 1px; font-size: .75rem;
                  color: var(--text); font-weight: 700; }}
    footer {{ text-align: center; color: var(--muted); font-size: .75rem;
              margin-top: 16px; }}
  </style>
</head>
<body>
<h1>📊 Dashboard de Lote — DWG Quality Checker</h1>
<p class="sub">Gerado em {ts}  ·  {total} arquivo(s) verificado(s)</p>

<div class="cards">
  <div class="card">
    <div class="val" style="color:var(--accent)">{rate:.0f}%</div>
    <div class="lab">Taxa de Aprovação</div>
  </div>
  <div class="card">
    <div class="val" style="color:var(--success)">{passed}</div>
    <div class="lab">Aprovados ✅</div>
  </div>
  <div class="card">
    <div class="val" style="color:var(--error)">{failed}</div>
    <div class="lab">Reprovados ❌</div>
  </div>
  <div class="card">
    <div class="val" style="color:var(--error)">{tot_err}</div>
    <div class="lab">Total de Erros</div>
  </div>
  <div class="card">
    <div class="val" style="color:var(--warning)">{tot_war}</div>
    <div class="lab">Total de Avisos</div>
  </div>
</div>

<section>
  <h2>📋 Resultado por Arquivo</h2>
  <table>
    <thead>
      <tr>
        <th>Arquivo</th><th>Status</th>
        <th>Erros</th><th>Avisos</th><th>Infos</th>
      </tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
</section>

{rules_section}

<footer>DWG Quality Checker v2.5.0  ·  Vantara Tech  ·  Luiz Q. Melo</footer>
</body>
</html>"""

    Path(output_path).write_text(html, encoding="utf-8")
    return str(output_path)
